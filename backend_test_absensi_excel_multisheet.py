#!/usr/bin/env python3
"""
Backend test for Absensi Report Excel Multi-Sheet Export
Tests the rewritten GET /api/absensi/report/export endpoint that produces 7 sheets:
Rekapitulasi, Identitas, Absensi, Jam Kerja, Stock Opname, Lembur, Verifikasi
"""
import requests
import json
import time
from datetime import datetime
from pymongo import MongoClient
import subprocess
import os
import tempfile

# Config
BASE_URL = "https://absensi-foundation.preview.emergentagent.com"
MONGO_URL = "mongodb://localhost:27017"
DB_NAME = "cycle_count"

# Test counters
total_tests = 0
passed_tests = 0

def test(name):
    global total_tests
    total_tests += 1
    print(f"\n{'='*80}")
    print(f"TEST {total_tests}: {name}")
    print('='*80)

def passed(msg=""):
    global passed_tests
    passed_tests += 1
    print(f"✅ PASSED: {msg}")

def failed(msg=""):
    print(f"❌ FAILED: {msg}")

def info(msg):
    print(f"ℹ️  {msg}")

def login(username, password):
    """Login and return token"""
    resp = requests.post(f"{BASE_URL}/api/auth/login", json={"username": username, "password": password})
    if resp.status_code != 200:
        failed(f"Login failed for {username}: {resp.status_code} {resp.text}")
        return None
    data = resp.json()
    token = data.get("token")
    user = data.get("user", {})
    info(f"Logged in as {username}, user_id: {user.get('id')}")
    return token, user

def get_mongo_db():
    """Get MongoDB database connection"""
    client = MongoClient(MONGO_URL)
    return client[DB_NAME]

def setup_test_data():
    """Set up test data using mongosh"""
    info("Setting up test data via mongosh...")
    
    # First, get Cindy's user_id
    db = get_mongo_db()
    cindy = db.employees.find_one({"username": "cindy"})
    if not cindy:
        failed("Cindy user not found in database")
        return None
    
    cindy_id = cindy.get("id")
    info(f"Found Cindy with user_id: {cindy_id}")
    
    # Delete existing test records
    result = db.absensi_records.delete_many({
        "user_id": cindy_id,
        "date": {"$in": ["2026-08-20", "2026-08-21", "2026-08-22"]}
    })
    info(f"Deleted {result.deleted_count} existing test records")
    
    # Insert test records
    test_records = [
        {
            "id": "T1",
            "user_id": cindy_id,
            "user_name": "Cindy",
            "user_role": "staff",
            "date": "2026-08-20",
            "shift_key": "apotek_pagi",
            "shift_name": "Apotek — Pagi",
            "shift_category": "apotek",
            "shift_start": "07:00",
            "shift_end": "15:00",
            "shift_start_mins": 420,
            "shift_end_mins": 900,
            "actual_check_in": datetime.fromisoformat("2026-08-20T00:00:00+00:00"),
            "actual_check_in_wita": "08:00",
            "effective_check_in_mins": 480,
            "late_minutes": 60,
            "actual_check_out": datetime.fromisoformat("2026-08-20T08:00:00+00:00"),
            "actual_check_out_wita": "16:00",
            "worked_minutes": 480,
            "check_in_lat": -8.65,
            "check_in_lng": 115.21,
            "check_in_distance_m": 20,
            "check_out_lat": -8.65,
            "check_out_lng": 115.21,
            "check_out_distance_m": 25,
            "so_selected": False,
            "overtime_requested": False,
            "overtime_minutes": 0,
            "overtime_status": "none",
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow()
        },
        {
            "id": "T2",
            "user_id": cindy_id,
            "user_name": "Cindy",
            "user_role": "staff",
            "date": "2026-08-21",
            "shift_key": "apotek_sore",
            "shift_name": "Apotek — Sore",
            "shift_category": "apotek",
            "shift_start": "15:00",
            "shift_end": "22:00",
            "shift_start_mins": 900,
            "shift_end_mins": 1320,
            "actual_check_in": datetime.fromisoformat("2026-08-21T06:00:00+00:00"),
            "actual_check_in_wita": "14:00",
            "effective_check_in_mins": 840,
            "late_minutes": 0,
            "actual_check_out": datetime.fromisoformat("2026-08-21T14:00:00+00:00"),
            "actual_check_out_wita": "22:00",
            "worked_minutes": 480,
            "check_in_lat": -8.65,
            "check_in_lng": 115.21,
            "check_in_distance_m": 15,
            "check_out_lat": -8.65,
            "check_out_lng": 115.21,
            "check_out_distance_m": 18,
            "so_selected": True,
            "so_effective_start_mins": 840,
            "overtime_requested": False,
            "overtime_minutes": 0,
            "overtime_status": "none",
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow()
        },
        {
            "id": "T3",
            "user_id": cindy_id,
            "user_name": "Cindy",
            "user_role": "staff",
            "date": "2026-08-22",
            "shift_key": "apotek_pagi",
            "shift_name": "Apotek — Pagi",
            "shift_category": "apotek",
            "shift_start": "07:00",
            "shift_end": "15:00",
            "shift_start_mins": 420,
            "shift_end_mins": 900,
            "actual_check_in": datetime.fromisoformat("2026-08-22T00:00:00+00:00"),
            "actual_check_in_wita": "08:00",
            "effective_check_in_mins": 480,
            "late_minutes": 60,
            "actual_check_out": datetime.fromisoformat("2026-08-22T09:00:00+00:00"),
            "actual_check_out_wita": "17:00",
            "worked_minutes": 540,
            "check_in_lat": -8.65,
            "check_in_lng": 115.21,
            "check_in_distance_m": 20,
            "check_out_lat": -8.65,
            "check_out_lng": 115.21,
            "check_out_distance_m": 25,
            "so_selected": False,
            "overtime_requested": True,
            "overtime_reason": "tutup toko",
            "overtime_raw_minutes": 120,
            "overtime_minutes": 120,
            "overtime_status": "approved",
            "overtime_reviewed_by_name": "Owner",
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow()
        }
    ]
    
    result = db.absensi_records.insert_many(test_records)
    info(f"Inserted {len(result.inserted_ids)} test records (T1, T2, T3)")
    
    return cindy_id

def cleanup_test_data(cindy_id):
    """Clean up test data"""
    info("Cleaning up test data...")
    db = get_mongo_db()
    result = db.absensi_records.delete_many({
        "user_id": cindy_id,
        "date": {"$in": ["2026-08-20", "2026-08-21", "2026-08-22"]}
    })
    info(f"Deleted {result.deleted_count} test records")

def parse_xlsx_with_openpyxl(filepath):
    """Parse XLSX file using openpyxl"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        return wb
    except ImportError:
        failed("openpyxl not installed. Installing...")
        subprocess.run(["pip", "install", "openpyxl"], check=True)
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        return wb

def main():
    print("\n" + "="*80)
    print("BACKEND TEST: Absensi Report Excel Multi-Sheet Export")
    print("="*80)
    
    # ========================================================================
    # SETUP: Login and prepare test data
    # ========================================================================
    test("SETUP: Login as owner and set up test data")
    
    owner_token, owner_user = login("owner", "owner123")
    if not owner_token:
        failed("Cannot proceed without owner token")
        return
    passed("Owner login successful")
    
    cindy_id = setup_test_data()
    if not cindy_id:
        failed("Cannot proceed without test data")
        return
    passed("Test data setup complete")
    
    # ========================================================================
    # TEST 1: GET /api/absensi/report/export returns XLSX
    # ========================================================================
    test("TEST 1: GET /api/absensi/report/export returns XLSX binary")
    
    headers = {"Authorization": f"Bearer {owner_token}"}
    resp = requests.get(
        f"{BASE_URL}/api/absensi/report/export?from=2026-08-01&to=2026-08-31",
        headers=headers
    )
    
    if resp.status_code != 200:
        failed(f"Expected 200, got {resp.status_code}: {resp.text}")
        cleanup_test_data(cindy_id)
        return
    passed(f"Status code: {resp.status_code}")
    
    content_type = resp.headers.get('Content-Type', '')
    if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' not in content_type:
        failed(f"Expected XLSX content-type, got: {content_type}")
        cleanup_test_data(cindy_id)
        return
    passed(f"Content-Type: {content_type}")
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        f.write(resp.content)
        xlsx_path = f.name
    info(f"Saved XLSX to: {xlsx_path}")
    passed(f"XLSX file size: {len(resp.content)} bytes")
    
    # ========================================================================
    # TEST 2: Parse XLSX and verify sheet order + names
    # ========================================================================
    test("TEST 2: Sheet order and names")
    
    wb = parse_xlsx_with_openpyxl(xlsx_path)
    sheet_names = wb.sheetnames
    expected_sheets = ['Rekapitulasi', 'Identitas', 'Absensi', 'Jam Kerja', 'Stock Opname', 'Lembur', 'Verifikasi']
    
    info(f"Found sheets: {sheet_names}")
    if sheet_names != expected_sheets:
        failed(f"Expected sheets: {expected_sheets}, got: {sheet_names}")
        cleanup_test_data(cindy_id)
        return
    passed(f"All 7 sheets present in correct order: {sheet_names}")
    
    # ========================================================================
    # TEST 3: Rekapitulasi sheet
    # ========================================================================
    test("TEST 3: Rekapitulasi sheet - headers and calculations")
    
    rekap_sheet = wb['Rekapitulasi']
    rekap_header = [cell.value for cell in rekap_sheet[1]]
    expected_rekap_header = [
        'Nama Staff',
        'Total Jam Kerja Diakui (jam)',
        'Total Jam SO Diakui (jam)',
        'Total Jam Lembur Diakui (jam)',
        'Total Jam Diakui (jam)'
    ]
    
    info(f"Rekapitulasi header: {rekap_header}")
    if rekap_header != expected_rekap_header:
        failed(f"Expected header: {expected_rekap_header}")
        cleanup_test_data(cindy_id)
        return
    passed("Rekapitulasi header correct")
    
    # Check Cindy's row (row 2)
    if rekap_sheet.max_row < 2:
        failed("No data rows in Rekapitulasi")
        cleanup_test_data(cindy_id)
        return
    
    cindy_row = [cell.value for cell in rekap_sheet[2]]
    info(f"Cindy's row: {cindy_row}")
    
    # Verify calculations:
    # T1: worked=480 (no OT) → 8h work
    # T2: worked=480 SO → 8h work + 8h SO
    # T3: worked=540, OT approved=120 → work=(540-120)/60=7h, OT=2h
    # Total work = 8+8+7 = 23h, SO = 8h, OT = 2h, Total = work+OT = 25h
    
    name = cindy_row[0]
    work_hours = cindy_row[1]
    so_hours = cindy_row[2]
    ot_hours = cindy_row[3]
    total_hours = cindy_row[4]
    
    if name != "Cindy":
        failed(f"Expected name 'Cindy', got: {name}")
        cleanup_test_data(cindy_id)
        return
    passed(f"Name: {name}")
    
    # Allow small rounding differences
    expected_work = 23.0  # (480-0)/60 + (480-0)/60 + (540-120)/60 = 8+8+7
    expected_so = 8.0     # T2 only: (480-0)/60
    expected_ot = 2.0     # T3 only: 120/60
    expected_total = 25.0 # work + ot = 23 + 2
    
    if abs(work_hours - expected_work) > 0.5:
        failed(f"Work hours: expected ~{expected_work}, got {work_hours}")
    else:
        passed(f"Work hours: {work_hours} (expected ~{expected_work})")
    
    if abs(so_hours - expected_so) > 0.5:
        failed(f"SO hours: expected ~{expected_so}, got {so_hours}")
    else:
        passed(f"SO hours: {so_hours} (expected ~{expected_so})")
    
    if abs(ot_hours - expected_ot) > 0.1:
        failed(f"OT hours: expected {expected_ot}, got {ot_hours}")
    else:
        passed(f"OT hours: {ot_hours} (expected {expected_ot})")
    
    if abs(total_hours - expected_total) > 0.5:
        failed(f"Total hours: expected ~{expected_total}, got {total_hours}")
    else:
        passed(f"Total hours: {total_hours} (expected ~{expected_total})")
    
    # Verify sum equation: total ≈ work + ot
    if abs(total_hours - (work_hours + ot_hours)) > 0.1:
        failed(f"Total ({total_hours}) != Work ({work_hours}) + OT ({ot_hours})")
    else:
        passed(f"Sum equation verified: {total_hours} ≈ {work_hours} + {ot_hours}")
    
    # ========================================================================
    # TEST 4: Identitas sheet
    # ========================================================================
    test("TEST 4: Identitas sheet - headers and row count")
    
    identitas_sheet = wb['Identitas']
    identitas_header = [cell.value for cell in identitas_sheet[1]]
    expected_identitas_header = ['Tanggal', 'Nama Staff', 'Role/Bagian', 'Shift', 'Jadwal Shift']
    
    info(f"Identitas header: {identitas_header}")
    if identitas_header != expected_identitas_header:
        failed(f"Expected header: {expected_identitas_header}")
        cleanup_test_data(cindy_id)
        return
    passed("Identitas header correct")
    
    # Should have 3 data rows (T1, T2, T3)
    data_rows = identitas_sheet.max_row - 1
    if data_rows != 3:
        failed(f"Expected 3 data rows, got {data_rows}")
    else:
        passed(f"Row count: {data_rows} (expected 3)")
    
    # ========================================================================
    # TEST 5: Absensi sheet
    # ========================================================================
    test("TEST 5: Absensi sheet - headers and status text")
    
    absensi_sheet = wb['Absensi']
    absensi_header = [cell.value for cell in absensi_sheet[1]]
    expected_absensi_header = [
        'Tanggal', 'Nama Staff', 'Shift', 'Jadwal Shift',
        'Jam Masuk', 'Jam Keluar', 'Status Kehadiran', 'Menit Terlambat'
    ]
    
    info(f"Absensi header: {absensi_header}")
    if absensi_header != expected_absensi_header:
        failed(f"Expected header: {expected_absensi_header}")
        cleanup_test_data(cindy_id)
        return
    passed("Absensi header correct")
    
    # Check status text for T1 (late) and T2 (on time)
    t1_status = absensi_sheet.cell(row=2, column=7).value  # First data row
    t2_status = absensi_sheet.cell(row=3, column=7).value  # Second data row
    
    info(f"T1 status: {t1_status}, T2 status: {t2_status}")
    
    # T1 has late_minutes=60, should be "Terlambat"
    # T2 has late_minutes=0, should be "Tepat Waktu"
    if 'Terlambat' not in str(t1_status):
        failed(f"T1 should be 'Terlambat', got: {t1_status}")
    else:
        passed(f"T1 status: {t1_status}")
    
    if 'Tepat Waktu' not in str(t2_status):
        failed(f"T2 should be 'Tepat Waktu', got: {t2_status}")
    else:
        passed(f"T2 status: {t2_status}")
    
    # ========================================================================
    # TEST 6: Jam Kerja sheet
    # ========================================================================
    test("TEST 6: Jam Kerja sheet - headers and calculations")
    
    jam_kerja_sheet = wb['Jam Kerja']
    jam_kerja_header = [cell.value for cell in jam_kerja_sheet[1]]
    expected_jam_kerja_header = [
        'Tanggal', 'Nama Staff',
        'Jam Kerja Normal (jam)', 'Jam Kerja Aktual (jam)', 'Jam Kerja Diakui (jam)'
    ]
    
    info(f"Jam Kerja header: {jam_kerja_header}")
    if jam_kerja_header != expected_jam_kerja_header:
        failed(f"Expected header: {expected_jam_kerja_header}")
        cleanup_test_data(cindy_id)
        return
    passed("Jam Kerja header correct")
    
    # Check T1: Normal=8.0, Aktual should match worked_minutes, Diakui=8.0
    # Note: Data is sorted by date DESC, so T3 (2026-08-22) is row 2, T2 (2026-08-21) is row 3, T1 (2026-08-20) is row 4
    t1_normal = jam_kerja_sheet.cell(row=4, column=3).value  # T1 is row 4 (sorted desc)
    t1_aktual = jam_kerja_sheet.cell(row=4, column=4).value
    t1_diakui = jam_kerja_sheet.cell(row=4, column=5).value
    
    info(f"T1 Jam Kerja: Normal={t1_normal}, Aktual={t1_aktual}, Diakui={t1_diakui}")
    
    if abs(t1_normal - 8.0) > 0.1:
        failed(f"T1 Normal: expected 8.0, got {t1_normal}")
    else:
        passed(f"T1 Normal: {t1_normal}")
    
    # Aktual is calculated from actual timestamps, Diakui is from worked_minutes field
    # Both should be 8.0 for T1 (worked_minutes=480)
    if abs(t1_diakui - 8.0) > 0.1:
        failed(f"T1 Diakui: expected 8.0, got {t1_diakui}")
    else:
        passed(f"T1 Diakui: {t1_diakui}")
    
    # Aktual might differ slightly due to timestamp precision, just verify it's reasonable
    if t1_aktual is None or t1_aktual == '-':
        failed(f"T1 Aktual: expected numeric value, got {t1_aktual}")
    else:
        passed(f"T1 Aktual: {t1_aktual} (calculated from timestamps)")
    
    # ========================================================================
    # TEST 7: Stock Opname sheet
    # ========================================================================
    test("TEST 7: Stock Opname sheet - only SO records, effective SO time")
    
    so_sheet = wb['Stock Opname']
    so_header = [cell.value for cell in so_sheet[1]]
    expected_so_header = [
        'Tanggal', 'Nama Staff', 'Shift', 'Status SO',
        'Jam Masuk SO', 'Jam Kerja Efektif SO', 'Jam SO Diakui (jam)'
    ]
    
    info(f"Stock Opname header: {so_header}")
    if so_header != expected_so_header:
        failed(f"Expected header: {expected_so_header}")
        cleanup_test_data(cindy_id)
        return
    passed("Stock Opname header correct")
    
    # Should have only 1 data row (T2 with so_selected=true)
    so_data_rows = so_sheet.max_row - 1
    if so_data_rows != 1:
        failed(f"Expected 1 SO data row, got {so_data_rows}")
    else:
        passed(f"SO row count: {so_data_rows} (only T2)")
    
    # Check T2 SO details
    t2_status_so = so_sheet.cell(row=2, column=4).value
    t2_jam_masuk_so = so_sheet.cell(row=2, column=5).value
    t2_efektif_so = so_sheet.cell(row=2, column=6).value
    
    info(f"T2 SO: Status={t2_status_so}, Jam Masuk={t2_jam_masuk_so}, Efektif={t2_efektif_so}")
    
    if t2_status_so != 'Ya':
        failed(f"T2 Status SO: expected 'Ya', got {t2_status_so}")
    else:
        passed(f"T2 Status SO: {t2_status_so}")
    
    # Efektif SO should be like '14:00-22:00' (check-in to check-out)
    if '14:00' not in str(t2_efektif_so) or '22:00' not in str(t2_efektif_so):
        failed(f"T2 Efektif SO: expected '14:00-22:00', got {t2_efektif_so}")
    else:
        passed(f"T2 Efektif SO: {t2_efektif_so}")
    
    # ========================================================================
    # TEST 8: Lembur sheet
    # ========================================================================
    test("TEST 8: Lembur sheet - only approved OT, rejected/pending excluded")
    
    lembur_sheet = wb['Lembur']
    lembur_header = [cell.value for cell in lembur_sheet[1]]
    expected_lembur_header = [
        'Tanggal', 'Nama Staff', 'Shift', 'Jam Selesai Shift',
        'Jam Mulai Lembur', 'Jam Selesai Lembur', 'Potensi Lembur (jam)',
        'Alasan', 'Status Approval', 'Approver', 'Jam Lembur Diakui (jam)'
    ]
    
    info(f"Lembur header: {lembur_header}")
    if lembur_header != expected_lembur_header:
        failed(f"Expected header: {expected_lembur_header}")
        cleanup_test_data(cindy_id)
        return
    passed("Lembur header correct (11 columns)")
    
    # Should have only 1 data row (T3 with approved OT)
    lembur_data_rows = lembur_sheet.max_row - 1
    if lembur_data_rows != 1:
        failed(f"Expected 1 Lembur data row, got {lembur_data_rows}")
    else:
        passed(f"Lembur row count: {lembur_data_rows} (only T3 approved)")
    
    # Check T3 OT details
    t3_alasan = lembur_sheet.cell(row=2, column=8).value
    t3_status = lembur_sheet.cell(row=2, column=9).value
    t3_diakui = lembur_sheet.cell(row=2, column=11).value
    
    info(f"T3 Lembur: Alasan={t3_alasan}, Status={t3_status}, Diakui={t3_diakui}")
    
    if 'tutup toko' not in str(t3_alasan):
        failed(f"T3 Alasan: expected 'tutup toko', got {t3_alasan}")
    else:
        passed(f"T3 Alasan: {t3_alasan}")
    
    if t3_status != 'approved':
        failed(f"T3 Status: expected 'approved', got {t3_status}")
    else:
        passed(f"T3 Status: {t3_status}")
    
    if abs(t3_diakui - 2.0) > 0.1:
        failed(f"T3 Diakui: expected 2.0 hours, got {t3_diakui}")
    else:
        passed(f"T3 Diakui: {t3_diakui} hours")
    
    # ========================================================================
    # TEST 9: Verifikasi sheet
    # ========================================================================
    test("TEST 9: Verifikasi sheet - GPS status validation")
    
    verif_sheet = wb['Verifikasi']
    verif_header = [cell.value for cell in verif_sheet[1]]
    
    # Check that header includes GPS and photo columns
    header_str = ' '.join([str(h) for h in verif_header])
    required_cols = [
        'Status Foto Masuk', 'Latitude Masuk', 'Longitude Masuk',
        'Jarak Masuk', 'Radius Masuk', 'Status GPS Masuk',
        'Status Foto Keluar', 'Latitude Keluar', 'Longitude Keluar',
        'Jarak Keluar', 'Radius Keluar', 'Status GPS Keluar'
    ]
    
    info(f"Verifikasi header: {verif_header}")
    
    missing_cols = [col for col in required_cols if col not in header_str]
    if missing_cols:
        failed(f"Missing columns in Verifikasi: {missing_cols}")
        cleanup_test_data(cindy_id)
        return
    passed("Verifikasi header includes all GPS and photo columns")
    
    # Check GPS status for T1 (distance_m=20, radius=50 → Valid)
    # Find column indices
    col_indices = {cell.value: idx+1 for idx, cell in enumerate(verif_sheet[1])}
    gps_in_col = col_indices.get('Status GPS Masuk')
    gps_out_col = col_indices.get('Status GPS Keluar')
    
    t1_gps_in = verif_sheet.cell(row=2, column=gps_in_col).value
    t1_gps_out = verif_sheet.cell(row=2, column=gps_out_col).value
    
    info(f"T1 GPS: In={t1_gps_in}, Out={t1_gps_out}")
    
    # T1 has distance 20m and 25m, both <= 50m radius → Valid
    if t1_gps_in != 'Valid':
        failed(f"T1 GPS In: expected 'Valid', got {t1_gps_in}")
    else:
        passed(f"T1 GPS In: {t1_gps_in}")
    
    if t1_gps_out != 'Valid':
        failed(f"T1 GPS Out: expected 'Valid', got {t1_gps_out}")
    else:
        passed(f"T1 GPS Out: {t1_gps_out}")
    
    # ========================================================================
    # TEST 10: Filter consistency - with user_id filter
    # ========================================================================
    test("TEST 10: Filter consistency - with user_id filter")
    
    # Create a fake user_id that doesn't exist
    resp_empty = requests.get(
        f"{BASE_URL}/api/absensi/report/export?from=2026-08-01&to=2026-08-31&user_id=fake-user-id-12345",
        headers=headers
    )
    
    if resp_empty.status_code != 200:
        failed(f"Expected 200 for empty filter, got {resp_empty.status_code}")
    else:
        passed("Empty filter returns 200")
    
    # Save and parse
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        f.write(resp_empty.content)
        xlsx_empty_path = f.name
    
    wb_empty = parse_xlsx_with_openpyxl(xlsx_empty_path)
    
    # All sheets should still exist
    if wb_empty.sheetnames != expected_sheets:
        failed(f"Empty filter: sheets missing, got {wb_empty.sheetnames}")
    else:
        passed("Empty filter: all 7 sheets still exist")
    
    # Rekap should have 0 staff rows (only header)
    rekap_empty = wb_empty['Rekapitulasi']
    if rekap_empty.max_row != 1:
        failed(f"Empty Rekap: expected 1 row (header only), got {rekap_empty.max_row}")
    else:
        passed(f"Empty Rekap: {rekap_empty.max_row} row (header only)")
    
    # Identitas should have 0 data rows
    identitas_empty = wb_empty['Identitas']
    if identitas_empty.max_row != 1:
        failed(f"Empty Identitas: expected 1 row (header only), got {identitas_empty.max_row}")
    else:
        passed(f"Empty Identitas: {identitas_empty.max_row} row (header only)")
    
    # SO and Lembur should have placeholder row with '-'
    so_empty = wb_empty['Stock Opname']
    if so_empty.max_row < 2:
        failed(f"Empty SO: expected at least 2 rows (header + placeholder), got {so_empty.max_row}")
    else:
        passed(f"Empty SO: {so_empty.max_row} rows (header + placeholder)")
    
    os.unlink(xlsx_empty_path)
    
    # ========================================================================
    # TEST 11: Regression - JSON report still works
    # ========================================================================
    test("TEST 11: Regression - JSON report endpoint unchanged")
    
    resp_json = requests.get(
        f"{BASE_URL}/api/absensi/report?from=2026-08-01&to=2026-08-31",
        headers=headers
    )
    
    if resp_json.status_code != 200:
        failed(f"JSON report: expected 200, got {resp_json.status_code}")
    else:
        passed("JSON report: status 200")
    
    json_data = resp_json.json()
    if 'items' not in json_data or 'filter' not in json_data or 'total' not in json_data or 'location' not in json_data:
        failed(f"JSON report: missing expected fields, got keys: {json_data.keys()}")
    else:
        passed(f"JSON report: has items, filter, total, location fields")
    
    if len(json_data['items']) != 3:
        failed(f"JSON report: expected 3 items, got {len(json_data['items'])}")
    else:
        passed(f"JSON report: {len(json_data['items'])} items (T1, T2, T3)")
    
    # ========================================================================
    # TEST 12: Add rejected/pending OT records and verify Diakui = 0
    # ========================================================================
    test("TEST 12: Lembur rejected/pending - Jam Lembur Diakui = 0")
    
    db = get_mongo_db()
    
    # Add T4 with rejected OT
    t4_record = {
        "id": "T4",
        "user_id": cindy_id,
        "user_name": "Cindy",
        "user_role": "staff",
        "date": "2026-08-23",
        "shift_key": "apotek_pagi",
        "shift_name": "Apotek — Pagi",
        "shift_category": "apotek",
        "shift_start": "07:00",
        "shift_end": "15:00",
        "shift_start_mins": 420,
        "shift_end_mins": 900,
        "actual_check_in": datetime.fromisoformat("2026-08-23T00:00:00+00:00"),
        "actual_check_in_wita": "08:00",
        "effective_check_in_mins": 480,
        "late_minutes": 60,
        "actual_check_out": datetime.fromisoformat("2026-08-23T08:00:00+00:00"),
        "actual_check_out_wita": "16:00",
        "worked_minutes": 480,
        "check_in_lat": -8.65,
        "check_in_lng": 115.21,
        "check_in_distance_m": 20,
        "check_out_lat": -8.65,
        "check_out_lng": 115.21,
        "check_out_distance_m": 25,
        "so_selected": False,
        "overtime_requested": True,
        "overtime_reason": "test rejected",
        "overtime_raw_minutes": 60,
        "overtime_minutes": 0,  # Rejected → 0
        "overtime_status": "rejected",
        "overtime_reviewed_by_name": "Owner",
        "createdAt": datetime.utcnow(),
        "updatedAt": datetime.utcnow()
    }
    
    # Add T5 with pending OT
    t5_record = {
        "id": "T5",
        "user_id": cindy_id,
        "user_name": "Cindy",
        "user_role": "staff",
        "date": "2026-08-24",
        "shift_key": "apotek_pagi",
        "shift_name": "Apotek — Pagi",
        "shift_category": "apotek",
        "shift_start": "07:00",
        "shift_end": "15:00",
        "shift_start_mins": 420,
        "shift_end_mins": 900,
        "actual_check_in": datetime.fromisoformat("2026-08-24T00:00:00+00:00"),
        "actual_check_in_wita": "08:00",
        "effective_check_in_mins": 480,
        "late_minutes": 60,
        "actual_check_out": datetime.fromisoformat("2026-08-24T08:00:00+00:00"),
        "actual_check_out_wita": "16:00",
        "worked_minutes": 480,
        "check_in_lat": -8.65,
        "check_in_lng": 115.21,
        "check_in_distance_m": 20,
        "check_out_lat": -8.65,
        "check_out_lng": 115.21,
        "check_out_distance_m": 25,
        "so_selected": False,
        "overtime_requested": True,
        "overtime_reason": "test pending",
        "overtime_raw_minutes": 60,
        "overtime_minutes": 60,  # Pending → not yet approved
        "overtime_status": "pending",
        "createdAt": datetime.utcnow(),
        "updatedAt": datetime.utcnow()
    }
    
    db.absensi_records.insert_many([t4_record, t5_record])
    info("Inserted T4 (rejected) and T5 (pending)")
    
    # Fetch new export
    resp_with_rejected = requests.get(
        f"{BASE_URL}/api/absensi/report/export?from=2026-08-01&to=2026-08-31",
        headers=headers
    )
    
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        f.write(resp_with_rejected.content)
        xlsx_rejected_path = f.name
    
    wb_rejected = parse_xlsx_with_openpyxl(xlsx_rejected_path)
    lembur_rejected = wb_rejected['Lembur']
    
    # Should now have 3 rows: T5 (pending), T4 (rejected), T3 (approved) - sorted by date DESC
    lembur_rows = lembur_rejected.max_row - 1
    if lembur_rows != 3:
        failed(f"Expected 3 Lembur rows (T3, T4, T5), got {lembur_rows}")
    else:
        passed(f"Lembur rows: {lembur_rows} (T3 approved, T4 rejected, T5 pending)")
    
    # Find T4 and T5 rows and check Diakui column
    # Data is sorted by date DESC: T5 (2026-08-24) row 2, T4 (2026-08-23) row 3, T3 (2026-08-22) row 4
    diakui_col = 11  # 'Jam Lembur Diakui (jam)' is column 11
    status_col = 9   # 'Status Approval' is column 9
    
    # Check all 3 rows and find which is which by status
    rows_data = []
    for row_idx in range(2, 5):  # rows 2, 3, 4
        status = lembur_rejected.cell(row=row_idx, column=status_col).value
        diakui = lembur_rejected.cell(row=row_idx, column=diakui_col).value
        rows_data.append((row_idx, status, diakui))
    
    info(f"Lembur rows data: {rows_data}")
    
    # Find rejected and pending rows
    rejected_row = next((r for r in rows_data if r[1] == 'rejected'), None)
    pending_row = next((r for r in rows_data if r[1] == 'pending'), None)
    approved_row = next((r for r in rows_data if r[1] == 'approved'), None)
    
    if rejected_row:
        if rejected_row[2] != 0:
            failed(f"Rejected row Diakui: expected 0, got {rejected_row[2]}")
        else:
            passed(f"Rejected (T4) Diakui: {rejected_row[2]}")
    else:
        failed("Rejected row not found")
    
    if pending_row:
        if pending_row[2] != 0:
            failed(f"Pending row Diakui: expected 0, got {pending_row[2]}")
        else:
            passed(f"Pending (T5) Diakui: {pending_row[2]}")
    else:
        failed("Pending row not found")
    
    if approved_row:
        if abs(approved_row[2] - 2.0) > 0.1:
            failed(f"Approved row Diakui: expected 2.0, got {approved_row[2]}")
        else:
            passed(f"Approved (T3) Diakui: {approved_row[2]}")
    else:
        failed("Approved row not found")
    
    # Clean up T4 and T5
    db.absensi_records.delete_many({"id": {"$in": ["T4", "T5"]}})
    os.unlink(xlsx_rejected_path)
    
    # ========================================================================
    # CLEANUP
    # ========================================================================
    test("CLEANUP: Delete test records")
    
    cleanup_test_data(cindy_id)
    os.unlink(xlsx_path)
    passed("Test data cleaned up")
    
    # ========================================================================
    # SUMMARY
    # ========================================================================
    print("\n" + "="*80)
    print(f"TEST SUMMARY: {passed_tests}/{total_tests} tests passed")
    print("="*80)
    
    if passed_tests == total_tests:
        print("✅ ALL TESTS PASSED - Absensi Excel Multi-Sheet Export FULLY WORKING")
    else:
        print(f"❌ {total_tests - passed_tests} TESTS FAILED")
    
    print("\n")

if __name__ == "__main__":
    main()
