#!/usr/bin/env python3
"""
Backend test: Jam Kerja Diakui logic verification for Absensi Excel export.

Verifies the deriveDiakui helper function in /app/lib/modules/absensi/service.js
by seeding 4 specific records and checking the Excel export output.

Test cases (all for user 'cindy'):
- V1: Aktual 8.68h (8j41m), Diakui 5.77h (5j46m) — late 134m, worked 521m
- V2: Aktual 6.45h (6j27m), Diakui 0.00h (0m) — late 515m, worked 387m
- V3: Aktual 8.17h (8j10m), Diakui 7.85h (7j51m) — late 9m, worked 490m
- V4: Aktual 6.88h (6j53m), Diakui 6.83h (6j50m) — late 10m, worked 413m

Rekapitulasi: Total Kerja Diakui = 5.77 + 0 + 7.85 + 6.83 = 20.45 (±0.05)
"""

import requests
import sys
from io import BytesIO
import openpyxl

BASE_URL = "https://absensi-foundation.preview.emergentagent.com"

def main():
    print("=" * 80)
    print("BACKEND TEST: Jam Kerja Diakui Logic (Absensi Excel Export)")
    print("=" * 80)
    
    # Step 1: Login as owner
    print("\n[1/7] Login as owner...")
    try:
        r = requests.post(f"{BASE_URL}/api/auth/login", json={"username": "owner", "password": "owner123"}, timeout=10)
        if r.status_code != 200:
            print(f"❌ FAIL: Owner login failed with status {r.status_code}")
            print(f"Response: {r.text}")
            return False
        token = r.json().get("token")
        if not token:
            print("❌ FAIL: No token in login response")
            return False
        headers = {"Authorization": f"Bearer {token}"}
        print(f"✅ PASS: Owner login successful (token: {token[:20]}...)")
    except Exception as e:
        print(f"❌ FAIL: Owner login exception: {e}")
        return False
    
    # Step 2: Get Cindy's user_id
    print("\n[2/7] Get Cindy's user_id from employees...")
    try:
        r = requests.get(f"{BASE_URL}/api/employees", headers=headers, timeout=10)
        if r.status_code != 200:
            print(f"❌ FAIL: GET /api/employees failed with status {r.status_code}")
            return False
        employees = r.json().get("items", [])
        cindy = next((e for e in employees if e.get("username") == "cindy"), None)
        if not cindy:
            print("❌ FAIL: User 'cindy' not found in employees")
            return False
        cindy_id = cindy.get("id")
        print(f"✅ PASS: Found Cindy (id: {cindy_id}, name: {cindy.get('name')})")
    except Exception as e:
        print(f"❌ FAIL: Get employees exception: {e}")
        return False
    
    # Step 3: Seed 4 test records via MongoDB
    print("\n[3/7] Seed 4 test records for Cindy via MongoDB...")
    try:
        import subprocess
        mongo_script = f"""
const uid = "{cindy_id}";
db.absensi_records.deleteMany({{id:{{$in:["V1","V2","V3","V4"]}}}});
const base={{user_id:uid,user_name:"Cindy",user_role:"staff",shift_category:"apotek",createdAt:new Date(),updatedAt:new Date(),
  so_selected:false,overtime_requested:false,overtime_status:"none",overtime_minutes:0}};
db.absensi_records.insertMany([
  {{ ...base,id:"V1",date:"2026-09-01",shift_key:"apotek_pagi",shift_name:"Apotek — Pagi",
    shift_start:"07:00",shift_end:"15:00",shift_start_mins:420,shift_end_mins:900,
    actual_check_in:new Date(),actual_check_in_wita:"09:14",
    actual_check_out:new Date(),actual_check_out_wita:"17:55",
    effective_check_in_mins:554,late_minutes:134,worked_minutes:521 }},
  {{ ...base,id:"V2",date:"2026-09-02",shift_key:"apotek_pagi",shift_name:"Apotek — Pagi",
    shift_start:"07:00",shift_end:"15:00",shift_start_mins:420,shift_end_mins:900,
    actual_check_in:new Date(),actual_check_in_wita:"15:35",
    actual_check_out:new Date(),actual_check_out_wita:"22:02",
    effective_check_in_mins:935,late_minutes:515,worked_minutes:387 }},
  {{ ...base,id:"V3",date:"2026-09-03",shift_key:"apotek_pagi",shift_name:"Apotek — Pagi",
    shift_start:"07:00",shift_end:"15:00",shift_start_mins:420,shift_end_mins:900,
    actual_check_in:new Date(),actual_check_in_wita:"07:09",
    actual_check_out:new Date(),actual_check_out_wita:"15:19",
    effective_check_in_mins:429,late_minutes:9,worked_minutes:490 }},
  {{ ...base,id:"V4",date:"2026-09-04",shift_key:"apotek_sore",shift_name:"Apotek — Sore",
    shift_start:"15:00",shift_end:"22:00",shift_start_mins:900,shift_end_mins:1320,
    actual_check_in:new Date(),actual_check_in_wita:"15:10",
    actual_check_out:new Date(),actual_check_out_wita:"22:03",
    effective_check_in_mins:910,late_minutes:10,worked_minutes:413 }}
]);
print("Inserted 4 records");
"""
        result = subprocess.run(
            ["mongosh", "mongodb://localhost:27017/cycle_count", "--quiet", "--eval", mongo_script],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            print(f"❌ FAIL: MongoDB seed failed: {result.stderr}")
            return False
        print(f"✅ PASS: Seeded 4 records (V1-V4) for Cindy")
        print(f"   MongoDB output: {result.stdout.strip()}")
    except Exception as e:
        print(f"❌ FAIL: MongoDB seed exception: {e}")
        return False
    
    # Step 4: Export Excel report
    print("\n[4/7] Export Excel report for 2026-09-01 to 2026-09-04...")
    try:
        r = requests.get(
            f"{BASE_URL}/api/absensi/report/export?from=2026-09-01&to=2026-09-04",
            headers=headers,
            timeout=30
        )
        if r.status_code != 200:
            print(f"❌ FAIL: Export failed with status {r.status_code}")
            print(f"Response: {r.text[:500]}")
            return False
        if not r.content or len(r.content) < 1000:
            print(f"❌ FAIL: Export returned empty or too small file ({len(r.content)} bytes)")
            return False
        print(f"✅ PASS: Export successful ({len(r.content)} bytes)")
    except Exception as e:
        print(f"❌ FAIL: Export exception: {e}")
        return False
    
    # Step 5: Parse Excel and verify "Jam Kerja" sheet
    print("\n[5/7] Parse Excel and verify 'Jam Kerja' sheet...")
    try:
        wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)
        if "Jam Kerja" not in wb.sheetnames:
            print(f"❌ FAIL: Sheet 'Jam Kerja' not found. Available: {wb.sheetnames}")
            return False
        ws = wb["Jam Kerja"]
        print(f"✅ PASS: Sheet 'Jam Kerja' found")
        
        # Expected values (hours, tolerance ±0.02)
        expected = {
            "V1": {"aktual": 8.68, "diakui": 5.77},
            "V2": {"aktual": 6.45, "diakui": 0.00},
            "V3": {"aktual": 8.17, "diakui": 7.85},
            "V4": {"aktual": 6.88, "diakui": 6.83},
        }
        
        # Find rows for V1-V4 (date column is A, name is B, aktual is D, diakui is E)
        results = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            date = str(row[0]) if row[0] else ""
            name = str(row[1]) if row[1] else ""
            aktual = row[3]  # Column D (0-indexed: 3)
            diakui = row[4]  # Column E (0-indexed: 4)
            
            # Match by date (2026-09-01 to 2026-09-04) and name (Cindy)
            if "2026-09-01" in date and "Cindy" in name:
                results["V1"] = {"aktual": aktual, "diakui": diakui}
            elif "2026-09-02" in date and "Cindy" in name:
                results["V2"] = {"aktual": aktual, "diakui": diakui}
            elif "2026-09-03" in date and "Cindy" in name:
                results["V3"] = {"aktual": aktual, "diakui": diakui}
            elif "2026-09-04" in date and "Cindy" in name:
                results["V4"] = {"aktual": aktual, "diakui": diakui}
        
        if len(results) != 4:
            print(f"❌ FAIL: Expected 4 rows for V1-V4, found {len(results)}")
            print(f"   Found: {list(results.keys())}")
            return False
        
        print(f"✅ PASS: Found all 4 rows (V1-V4)")
        
        # Verify each row
        all_pass = True
        for rec_id, exp in expected.items():
            actual_val = results[rec_id]
            
            # Convert to float (handle '-' or None)
            try:
                aktual_hours = float(actual_val["aktual"]) if actual_val["aktual"] not in [None, "-", ""] else 0.0
                diakui_hours = float(actual_val["diakui"]) if actual_val["diakui"] not in [None, "-", ""] else 0.0
            except (ValueError, TypeError) as e:
                print(f"   ❌ FAIL {rec_id}: Cannot parse values (aktual={actual_val['aktual']}, diakui={actual_val['diakui']})")
                all_pass = False
                continue
            
            # Check tolerance ±0.02
            aktual_ok = abs(aktual_hours - exp["aktual"]) <= 0.02
            diakui_ok = abs(diakui_hours - exp["diakui"]) <= 0.02
            
            if aktual_ok and diakui_ok:
                print(f"   ✅ PASS {rec_id}: Aktual={aktual_hours:.2f} (exp {exp['aktual']}), Diakui={diakui_hours:.2f} (exp {exp['diakui']})")
            else:
                print(f"   ❌ FAIL {rec_id}:")
                if not aktual_ok:
                    print(f"      Aktual: got {aktual_hours:.2f}, expected {exp['aktual']} (diff {abs(aktual_hours - exp['aktual']):.4f})")
                if not diakui_ok:
                    print(f"      Diakui: got {diakui_hours:.2f}, expected {exp['diakui']} (diff {abs(diakui_hours - exp['diakui']):.4f})")
                all_pass = False
        
        if not all_pass:
            return False
        
    except Exception as e:
        print(f"❌ FAIL: Excel parsing exception: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Step 6: Verify "Rekapitulasi" sheet
    print("\n[6/7] Verify 'Rekapitulasi' sheet...")
    try:
        if "Rekapitulasi" not in wb.sheetnames:
            print(f"❌ FAIL: Sheet 'Rekapitulasi' not found")
            return False
        ws_rekap = wb["Rekapitulasi"]
        print(f"✅ PASS: Sheet 'Rekapitulasi' found")
        
        # Find Cindy's row (name in column A, Total Jam Diakui in column E)
        cindy_row = None
        for row in ws_rekap.iter_rows(min_row=2, values_only=True):
            name = str(row[0]) if row[0] else ""
            if "Cindy" in name:
                cindy_row = row
                break
        
        if not cindy_row:
            print(f"❌ FAIL: Cindy not found in Rekapitulasi sheet")
            return False
        
        # Column E (0-indexed: 4) = Total Jam Diakui
        total_diakui = cindy_row[4]
        try:
            total_diakui_val = float(total_diakui) if total_diakui not in [None, "-", ""] else 0.0
        except (ValueError, TypeError):
            print(f"❌ FAIL: Cannot parse Total Jam Diakui: {total_diakui}")
            return False
        
        expected_total = 20.45
        if abs(total_diakui_val - expected_total) <= 0.05:
            print(f"✅ PASS: Cindy's Total Jam Diakui = {total_diakui_val:.2f} (expected {expected_total}, tolerance ±0.05)")
        else:
            print(f"❌ FAIL: Cindy's Total Jam Diakui = {total_diakui_val:.2f}, expected {expected_total} (diff {abs(total_diakui_val - expected_total):.4f})")
            return False
        
    except Exception as e:
        print(f"❌ FAIL: Rekapitulasi verification exception: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Step 7: Cleanup
    print("\n[7/7] Cleanup test records...")
    try:
        mongo_cleanup = """
db.absensi_records.deleteMany({id:{$in:["V1","V2","V3","V4"]}});
print("Deleted V1-V4");
"""
        result = subprocess.run(
            ["mongosh", "mongodb://localhost:27017/cycle_count", "--quiet", "--eval", mongo_cleanup],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            print(f"⚠️  WARNING: Cleanup failed: {result.stderr}")
        else:
            print(f"✅ PASS: Cleanup successful")
    except Exception as e:
        print(f"⚠️  WARNING: Cleanup exception: {e}")
    
    print("\n" + "=" * 80)
    print("✅ ALL TESTS PASSED — Jam Kerja Diakui logic is CORRECT")
    print("=" * 80)
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
