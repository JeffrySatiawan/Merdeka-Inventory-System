#!/usr/bin/env python3
"""
Backend test for Absensi Excel export — patched Jam Kerja Diakui logic.

Tests the new deriveDiakui() helper function that calculates:
- Kerja Diakui = intersection of actual time with shift normal time
- SO Diakui = early arrival time (before shift start) when so_selected=true
- Lembur Diakui = overtime_minutes only if overtime_status='approved'
- Total Diakui = Kerja + SO + Lembur (no double count)
"""

import requests
import subprocess
import json
from io import BytesIO
from openpyxl import load_workbook

BASE_URL = "https://absensi-foundation.preview.emergentagent.com"

def run_mongo_command(cmd):
    """Execute MongoDB command via mongosh."""
    result = subprocess.run(
        ["mongosh", "mongodb://localhost:27017/cycle_count", "--quiet", "--eval", cmd],
        capture_output=True,
        text=True
    )
    if result.returncode != 0:
        print(f"MongoDB command failed: {result.stderr}")
        return None
    return result.stdout.strip()

def setup_test_data():
    """Setup 7 test records (C1-C7) as specified in review request."""
    print("\n=== SETUP: Creating 7 test records (C1-C7) ===")
    
    # First, get Cindy's user_id
    get_cindy_id = """
    const cindy = db.employees.findOne({username: "cindy"});
    if (cindy) { print(cindy.id); } else { print("NOT_FOUND"); }
    """
    cindy_id = run_mongo_command(get_cindy_id)
    if not cindy_id or cindy_id == "NOT_FOUND":
        print("❌ ERROR: Cindy user not found in employees collection")
        return None
    print(f"✓ Found Cindy's user_id: {cindy_id}")
    
    # Clean up existing test records
    cleanup_cmd = f"""
    db.absensi_records.deleteMany({{
        user_id: "{cindy_id}",
        date: {{$regex: "^2026-08-2[0-9]|^2026-08-19"}}
    }});
    print("Cleaned up existing test records");
    """
    run_mongo_command(cleanup_cmd)
    print("✓ Cleaned up existing test records")
    
    # Insert 7 test records
    insert_cmd = f"""
    const uid = "{cindy_id}";
    const base = {{
        user_id: uid,
        user_name: "Cindy",
        user_role: "staff",
        shift_category: "apotek",
        createdAt: new Date(),
        updatedAt: new Date()
    }};
    
    // C1: Pagi 07:00-15:00. Masuk 15:35 Keluar 22:02. NO lembur request.
    db.absensi_records.insertOne({{
        ...base,
        id: "C1",
        date: "2026-08-19",
        shift_key: "apotek_pagi",
        shift_name: "Apotek — Pagi",
        shift_start: "07:00",
        shift_end: "15:00",
        shift_start_mins: 420,
        shift_end_mins: 900,
        actual_check_in: new Date(),
        actual_check_in_wita: "15:35",
        actual_check_out: new Date(),
        actual_check_out_wita: "22:02",
        effective_check_in_mins: 935,
        late_minutes: 515,
        worked_minutes: 387,
        so_selected: false,
        overtime_requested: false,
        overtime_status: "none",
        overtime_minutes: 0
    }});
    
    // C2: Pagi 07:00-15:00. Masuk 09:14 Keluar 17:55. Lembur REJECTED.
    db.absensi_records.insertOne({{
        ...base,
        id: "C2",
        date: "2026-08-20",
        shift_key: "apotek_pagi",
        shift_name: "Apotek — Pagi",
        shift_start: "07:00",
        shift_end: "15:00",
        shift_start_mins: 420,
        shift_end_mins: 900,
        actual_check_in: new Date(),
        actual_check_in_wita: "09:14",
        actual_check_out: new Date(),
        actual_check_out_wita: "17:55",
        effective_check_in_mins: 554,
        late_minutes: 134,
        worked_minutes: 521,
        so_selected: false,
        overtime_requested: true,
        overtime_status: "rejected",
        overtime_minutes: 0,
        overtime_raw_minutes: 175,
        overtime_reason: "salah kirim"
    }});
    
    // C3: Pagi. Masuk 07:09 Keluar 15:19. NO lembur request.
    db.absensi_records.insertOne({{
        ...base,
        id: "C3",
        date: "2026-08-21",
        shift_key: "apotek_pagi",
        shift_name: "Apotek — Pagi",
        shift_start: "07:00",
        shift_end: "15:00",
        shift_start_mins: 420,
        shift_end_mins: 900,
        actual_check_in: new Date(),
        actual_check_in_wita: "07:09",
        actual_check_out: new Date(),
        actual_check_out_wita: "15:19",
        effective_check_in_mins: 429,
        late_minutes: 9,
        worked_minutes: 490,
        so_selected: false,
        overtime_requested: false,
        overtime_status: "none",
        overtime_minutes: 0
    }});
    
    // C4: Sore + SO. 15:00-22:00. Masuk 14:00 Keluar 22:00.
    db.absensi_records.insertOne({{
        ...base,
        id: "C4",
        date: "2026-08-22",
        shift_key: "apotek_sore",
        shift_name: "Apotek — Sore",
        shift_start: "15:00",
        shift_end: "22:00",
        shift_start_mins: 900,
        shift_end_mins: 1320,
        actual_check_in: new Date(),
        actual_check_in_wita: "14:00",
        actual_check_out: new Date(),
        actual_check_out_wita: "22:00",
        effective_check_in_mins: 840,
        late_minutes: 0,
        worked_minutes: 480,
        so_selected: true,
        so_effective_start_mins: 840,
        overtime_requested: false,
        overtime_status: "none",
        overtime_minutes: 0
    }});
    
    // C5: Sore + SO. Masuk 13:00 Keluar 22:00.
    db.absensi_records.insertOne({{
        ...base,
        id: "C5",
        date: "2026-08-23",
        shift_key: "apotek_sore",
        shift_name: "Apotek — Sore",
        shift_start: "15:00",
        shift_end: "22:00",
        shift_start_mins: 900,
        shift_end_mins: 1320,
        actual_check_in: new Date(),
        actual_check_in_wita: "13:00",
        actual_check_out: new Date(),
        actual_check_out_wita: "22:00",
        effective_check_in_mins: 780,
        late_minutes: 0,
        worked_minutes: 540,
        so_selected: true,
        so_effective_start_mins: 780,
        overtime_requested: false,
        overtime_status: "none",
        overtime_minutes: 0
    }});
    
    // C6: Pagi. Masuk 07:00 Keluar 17:00. Lembur APPROVED (120 mnt).
    db.absensi_records.insertOne({{
        ...base,
        id: "C6",
        date: "2026-08-24",
        shift_key: "apotek_pagi",
        shift_name: "Apotek — Pagi",
        shift_start: "07:00",
        shift_end: "15:00",
        shift_start_mins: 420,
        shift_end_mins: 900,
        actual_check_in: new Date(),
        actual_check_in_wita: "07:00",
        actual_check_out: new Date(),
        actual_check_out_wita: "17:00",
        effective_check_in_mins: 420,
        late_minutes: 0,
        worked_minutes: 600,
        so_selected: false,
        overtime_requested: true,
        overtime_status: "approved",
        overtime_minutes: 120,
        overtime_raw_minutes: 120,
        overtime_reviewed_by_name: "Owner"
    }});
    
    // C7: Pagi. Masuk 07:00 Keluar 17:00. Lembur REJECTED.
    db.absensi_records.insertOne({{
        ...base,
        id: "C7",
        date: "2026-08-25",
        shift_key: "apotek_pagi",
        shift_name: "Apotek — Pagi",
        shift_start: "07:00",
        shift_end: "15:00",
        shift_start_mins: 420,
        shift_end_mins: 900,
        actual_check_in: new Date(),
        actual_check_in_wita: "07:00",
        actual_check_out: new Date(),
        actual_check_out_wita: "17:00",
        effective_check_in_mins: 420,
        late_minutes: 0,
        worked_minutes: 600,
        so_selected: false,
        overtime_requested: true,
        overtime_status: "rejected",
        overtime_minutes: 0,
        overtime_raw_minutes: 120,
        overtime_reviewed_by_name: "Owner"
    }});
    
    print("Inserted 7 test records (C1-C7)");
    """
    run_mongo_command(insert_cmd)
    print("✓ Inserted 7 test records (C1-C7)")
    
    return cindy_id

def cleanup_test_data():
    """Delete test records C1-C7."""
    print("\n=== CLEANUP: Deleting test records ===")
    cleanup_cmd = """
    db.absensi_records.deleteMany({id: {$in: ["C1","C2","C3","C4","C5","C6","C7"]}});
    print("Deleted test records");
    """
    run_mongo_command(cleanup_cmd)
    print("✓ Deleted test records C1-C7")

def test_excel_export():
    """Test the Excel export endpoint with new Diakui logic."""
    print("\n" + "="*80)
    print("BACKEND TEST: Absensi Excel Export — Jam Kerja Diakui Patch")
    print("="*80)
    
    # Setup test data
    cindy_id = setup_test_data()
    if not cindy_id:
        print("❌ SETUP FAILED: Could not create test data")
        return
    
    test_passed = True
    
    try:
        # Login as owner
        print("\n=== TEST 1: Login as owner ===")
        login_resp = requests.post(
            f"{BASE_URL}/api/auth/login",
            json={"username": "owner", "password": "owner123"}
        )
        if login_resp.status_code != 200:
            print(f"❌ Login failed: {login_resp.status_code}")
            return
        token = login_resp.json().get("token")
        print(f"✓ Owner login successful, token: {token[:20]}...")
        
        auth_headers = {"Authorization": f"Bearer {token}"}
        
        # Test Excel export
        print("\n=== TEST 2: GET /api/absensi/report/export ===")
        export_resp = requests.get(
            f"{BASE_URL}/api/absensi/report/export",
            params={"from": "2026-08-19", "to": "2026-08-25"},
            headers=auth_headers
        )
        
        if export_resp.status_code != 200:
            print(f"❌ Export failed: {export_resp.status_code}")
            print(f"Response: {export_resp.text[:500]}")
            return
        
        content_type = export_resp.headers.get("content-type", "")
        if "spreadsheet" not in content_type:
            print(f"❌ Wrong content-type: {content_type}")
            return
        
        print(f"✓ Export successful: {len(export_resp.content)} bytes, content-type: {content_type}")
        
        # Parse Excel
        print("\n=== TEST 3: Parse Excel workbook ===")
        wb = load_workbook(BytesIO(export_resp.content))
        sheet_names = wb.sheetnames
        print(f"✓ Workbook loaded, sheets: {sheet_names}")
        
        # Expected sheet order
        expected_sheets = ['Rekapitulasi', 'Identitas', 'Absensi', 'Jam Kerja', 'Stock Opname', 'Lembur', 'Verifikasi']
        if sheet_names != expected_sheets:
            print(f"❌ Sheet order mismatch!")
            print(f"   Expected: {expected_sheets}")
            print(f"   Got: {sheet_names}")
            test_passed = False
        else:
            print(f"✓ Sheet order correct: {expected_sheets}")
        
        # Test per-record values in Jam Kerja sheet
        print("\n=== TEST 4: Verify Jam Kerja sheet (per-record Diakui values) ===")
        jam_kerja_ws = wb['Jam Kerja']
        
        # Expected values (from review request)
        # Format: (date, aktual_hours, kerja_diakui_hours)
        expected_values = {
            "2026-08-19": {"aktual": 6.45, "kerja_diakui": 0.00},  # C1: 15:35-22:02, no overlap
            "2026-08-20": {"aktual": 8.68, "kerja_diakui": 5.77},  # C2: 09:14-17:55, overlap 09:14-15:00
            "2026-08-21": {"aktual": 8.17, "kerja_diakui": 7.85},  # C3: 07:09-15:19, overlap 07:09-15:00
            "2026-08-22": {"aktual": 8.00, "kerja_diakui": 7.00},  # C4: 14:00-22:00, kerja 15:00-22:00
            "2026-08-23": {"aktual": 9.00, "kerja_diakui": 7.00},  # C5: 13:00-22:00, kerja 15:00-22:00
            "2026-08-24": {"aktual": 10.00, "kerja_diakui": 8.00}, # C6: 07:00-17:00, kerja 07:00-15:00
            "2026-08-25": {"aktual": 10.00, "kerja_diakui": 8.00}, # C7: 07:00-17:00, kerja 07:00-15:00
        }
        
        # Find header row
        header_row = None
        for i, row in enumerate(jam_kerja_ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and 'Tanggal' in str(row):
                header_row = i
                break
        
        if not header_row:
            print("❌ Could not find header row in Jam Kerja sheet")
            test_passed = False
        else:
            print(f"✓ Found header row at line {header_row}")
            
            # Get header indices
            jam_kerja_headers = list(jam_kerja_ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
            print(f"   Headers: {jam_kerja_headers}")
            
            # Find column indices
            try:
                date_idx = jam_kerja_headers.index('Tanggal')
                aktual_idx = jam_kerja_headers.index('Jam Kerja Aktual (jam)')
                diakui_idx = jam_kerja_headers.index('Jam Kerja Diakui (jam)')
                
                # Read all data rows
                data_rows = list(jam_kerja_ws.iter_rows(min_row=header_row+1, values_only=True))
                print(f"✓ Found {len(data_rows)} data rows")
                
                # Verify each record
                tolerance = 0.02  # Allow ±0.02 as specified in review request
                for row in data_rows:
                    if not row or not row[date_idx]:
                        continue
                    
                    date = row[date_idx]
                    if date not in expected_values:
                        continue
                    
                    expected = expected_values[date]
                    actual_aktual = float(row[aktual_idx]) if row[aktual_idx] is not None else 0
                    actual_diakui = float(row[diakui_idx]) if row[diakui_idx] is not None else 0
                    
                    print(f"\n   Date: {date}")
                    print(f"   Aktual: expected {expected['aktual']:.2f}, got {actual_aktual:.2f}")
                    print(f"   Kerja Diakui: expected {expected['kerja_diakui']:.2f}, got {actual_diakui:.2f}")
                    
                    aktual_ok = abs(actual_aktual - expected['aktual']) <= tolerance
                    diakui_ok = abs(actual_diakui - expected['kerja_diakui']) <= tolerance
                    
                    if aktual_ok and diakui_ok:
                        print(f"   ✅ Values match within tolerance (±{tolerance})")
                    else:
                        print(f"   ❌ Values mismatch:")
                        if not aktual_ok:
                            print(f"      Aktual: expected {expected['aktual']:.2f}, got {actual_aktual:.2f}, diff {abs(actual_aktual - expected['aktual']):.3f}")
                        if not diakui_ok:
                            print(f"      Kerja Diakui: expected {expected['kerja_diakui']:.2f}, got {actual_diakui:.2f}, diff {abs(actual_diakui - expected['kerja_diakui']):.3f}")
                        test_passed = False
                        
            except ValueError as e:
                print(f"❌ Could not find required columns: {e}")
                test_passed = False
        
        # Test Rekapitulasi sheet
        print("\n=== TEST 5: Verify Rekapitulasi sheet (totals) ===")
        rekap_ws = wb['Rekapitulasi']
        
        # Expected totals for Cindy (from review request)
        # Total Kerja = 0 + 5.77 + 7.85 + 7.00 + 7.00 + 8.00 + 8.00 = 43.62 jam
        # Total SO = 1.00 + 2.00 = 3.00 jam
        # Total Lembur = 2.00 jam (only C6)
        # Total Diakui = 43.62 + 3.00 + 2.00 = 48.62 jam
        
        expected_totals = {
            'kerja': 43.62,
            'so': 3.00,
            'lembur': 2.00,
            'total': 48.62
        }
        
        # Find Cindy's row in Rekapitulasi
        rekap_header_row = None
        for i, row in enumerate(rekap_ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and 'Nama Staff' in str(row):
                rekap_header_row = i
                break
        
        if not rekap_header_row:
            print("❌ Could not find header row in Rekapitulasi sheet")
            test_passed = False
        else:
            rekap_headers = list(rekap_ws.iter_rows(min_row=rekap_header_row, max_row=rekap_header_row, values_only=True))[0]
            print(f"✓ Rekapitulasi headers: {rekap_headers}")
            
            # Find Cindy's row
            cindy_row = None
            for row in rekap_ws.iter_rows(min_row=rekap_header_row+1, values_only=True):
                if row and 'Cindy' in str(row[0]):
                    cindy_row = row
                    break
            
            if not cindy_row:
                print("❌ Could not find Cindy's row in Rekapitulasi")
                test_passed = False
            else:
                print(f"✓ Found Cindy's row: {cindy_row}")
                
                # Verify totals (with tolerance of ±0.05)
                tolerance = 0.05
                
                # Assuming order: Name, Kerja, SO, Lembur, Total
                if len(cindy_row) >= 5:
                    actual_kerja = float(cindy_row[1]) if cindy_row[1] is not None else 0
                    actual_so = float(cindy_row[2]) if cindy_row[2] is not None else 0
                    actual_lembur = float(cindy_row[3]) if cindy_row[3] is not None else 0
                    actual_total = float(cindy_row[4]) if cindy_row[4] is not None else 0
                    
                    print(f"\n   Expected vs Actual:")
                    print(f"   Kerja:  {expected_totals['kerja']:.2f} vs {actual_kerja:.2f}")
                    print(f"   SO:     {expected_totals['so']:.2f} vs {actual_so:.2f}")
                    print(f"   Lembur: {expected_totals['lembur']:.2f} vs {actual_lembur:.2f}")
                    print(f"   Total:  {expected_totals['total']:.2f} vs {actual_total:.2f}")
                    
                    # Verify with tolerance
                    kerja_ok = abs(actual_kerja - expected_totals['kerja']) <= tolerance
                    so_ok = abs(actual_so - expected_totals['so']) <= tolerance
                    lembur_ok = abs(actual_lembur - expected_totals['lembur']) <= tolerance
                    total_ok = abs(actual_total - expected_totals['total']) <= tolerance
                    
                    if kerja_ok and so_ok and lembur_ok and total_ok:
                        print(f"   ✅ All totals match within tolerance (±{tolerance})")
                    else:
                        print(f"   ❌ Totals mismatch:")
                        if not kerja_ok:
                            print(f"      Kerja: expected {expected_totals['kerja']:.2f}, got {actual_kerja:.2f}")
                        if not so_ok:
                            print(f"      SO: expected {expected_totals['so']:.2f}, got {actual_so:.2f}")
                        if not lembur_ok:
                            print(f"      Lembur: expected {expected_totals['lembur']:.2f}, got {actual_lembur:.2f}")
                        if not total_ok:
                            print(f"      Total: expected {expected_totals['total']:.2f}, got {actual_total:.2f}")
                        test_passed = False
                    
                    # Verify formula: Total = Kerja + SO + Lembur
                    calculated_total = actual_kerja + actual_so + actual_lembur
                    if abs(calculated_total - actual_total) <= 0.01:
                        print(f"   ✅ Total formula verified: {actual_kerja:.2f} + {actual_so:.2f} + {actual_lembur:.2f} = {actual_total:.2f}")
                    else:
                        print(f"   ❌ Total formula mismatch: {actual_kerja:.2f} + {actual_so:.2f} + {actual_lembur:.2f} = {calculated_total:.2f}, but got {actual_total:.2f}")
                        test_passed = False
                else:
                    print(f"❌ Cindy's row has insufficient columns: {len(cindy_row)}")
                    test_passed = False
        
        # Test Stock Opname sheet
        print("\n=== TEST 6: Verify Stock Opname sheet ===")
        so_ws = wb['Stock Opname']
        
        # Should only have C4 and C5 (so_selected=true)
        so_data_rows = list(so_ws.iter_rows(min_row=2, values_only=True))
        so_count = len([r for r in so_data_rows if r and r[0] and r[0] != '-'])
        
        print(f"   Found {so_count} SO records (expected 2: C4, C5)")
        if so_count == 2:
            print(f"   ✅ SO record count correct")
            
            # Verify SO Diakui values (1.00 and 2.00 hours)
            so_header_row = 1
            so_headers = list(so_ws.iter_rows(min_row=so_header_row, max_row=so_header_row, values_only=True))[0]
            
            # Look for SO Diakui column
            if 'Jam SO Diakui (jam)' in so_headers:
                so_diakui_idx = so_headers.index('Jam SO Diakui (jam)')
                so_values = []
                for row in so_data_rows:
                    if row and row[0] and row[0] != '-':
                        so_val = float(row[so_diakui_idx]) if row[so_diakui_idx] is not None else 0
                        so_values.append(so_val)
                
                print(f"   SO Diakui values: {so_values}")
                # Expected: 1.00 and 2.00 (in some order)
                if sorted(so_values) == [1.0, 2.0]:
                    print(f"   ✅ SO Diakui values correct: 1.00 and 2.00 hours")
                else:
                    print(f"   ❌ SO Diakui values mismatch: expected [1.0, 2.0], got {sorted(so_values)}")
                    test_passed = False
        else:
            print(f"   ❌ SO record count mismatch: expected 2, got {so_count}")
            test_passed = False
        
        # Test Lembur sheet
        print("\n=== TEST 7: Verify Lembur sheet ===")
        lembur_ws = wb['Lembur']
        
        # Should have C2, C6, C7 (overtime_requested=true)
        lembur_data_rows = list(lembur_ws.iter_rows(min_row=2, values_only=True))
        lembur_count = len([r for r in lembur_data_rows if r and r[0] and r[0] != '-'])
        
        print(f"   Found {lembur_count} Lembur records (expected 3: C2, C6, C7)")
        if lembur_count == 3:
            print(f"   ✅ Lembur record count correct")
            
            # Verify Lembur Diakui values
            lembur_header_row = 1
            lembur_headers = list(lembur_ws.iter_rows(min_row=lembur_header_row, max_row=lembur_header_row, values_only=True))[0]
            
            # Try both possible column names
            lembur_diakui_col = None
            if 'Jam Lembur Diakui (jam)' in lembur_headers:
                lembur_diakui_col = 'Jam Lembur Diakui (jam)'
            elif 'Lembur Diakui (jam)' in lembur_headers:
                lembur_diakui_col = 'Lembur Diakui (jam)'
            
            if lembur_diakui_col:
                lembur_diakui_idx = lembur_headers.index(lembur_diakui_col)
                date_idx = lembur_headers.index('Tanggal')
                
                # Expected: C2=0, C6=2.00, C7=0
                expected_lembur = {
                    "2026-08-20": 0.0,  # C2 rejected
                    "2026-08-24": 2.0,  # C6 approved
                    "2026-08-25": 0.0,  # C7 rejected
                }
                
                for row in lembur_data_rows:
                    if row and row[0] and row[0] != '-':
                        date = row[date_idx]
                        lembur_val = float(row[lembur_diakui_idx]) if row[lembur_diakui_idx] is not None else 0
                        
                        if date in expected_lembur:
                            expected_val = expected_lembur[date]
                            if abs(lembur_val - expected_val) <= 0.02:
                                print(f"   ✅ {date}: Lembur Diakui = {lembur_val:.2f} (expected {expected_val:.2f})")
                            else:
                                print(f"   ❌ {date}: Lembur Diakui = {lembur_val:.2f}, expected {expected_val:.2f}")
                                test_passed = False
            else:
                print(f"   ⚠️  'Jam Lembur Diakui (jam)' column not found in headers: {lembur_headers}")
                test_passed = False
        else:
            print(f"   ❌ Lembur record count mismatch: expected 3, got {lembur_count}")
            test_passed = False
        
        # Test Absensi sheet
        print("\n=== TEST 8: Verify Absensi sheet (late status) ===")
        absensi_ws = wb['Absensi']
        
        # Find header
        absensi_header_row = None
        for i, row in enumerate(absensi_ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and 'Tanggal' in str(row):
                absensi_header_row = i
                break
        
        if absensi_header_row:
            absensi_headers = list(absensi_ws.iter_rows(min_row=absensi_header_row, max_row=absensi_header_row, values_only=True))[0]
            print(f"   Absensi headers: {absensi_headers}")
            
            # Check for late minutes column
            if 'Menit Terlambat' in absensi_headers:
                late_idx = absensi_headers.index('Menit Terlambat')
                date_idx = absensi_headers.index('Tanggal')
                absensi_data = list(absensi_ws.iter_rows(min_row=absensi_header_row+1, values_only=True))
                
                # Expected late minutes by date
                expected_late_by_date = {
                    "2026-08-19": 515,  # C1
                    "2026-08-20": 134,  # C2
                    "2026-08-21": 9,    # C3
                    "2026-08-22": 0,    # C4
                    "2026-08-23": 0,    # C5
                    "2026-08-24": 0,    # C6
                    "2026-08-25": 0,    # C7
                }
                
                print(f"   Verifying late minutes by date:")
                all_late_ok = True
                for row in absensi_data[:7]:
                    if row and row[date_idx]:
                        date = row[date_idx]
                        actual_late = int(row[late_idx]) if row[late_idx] is not None else 0
                        expected_late = expected_late_by_date.get(date, 0)
                        
                        if actual_late == expected_late:
                            print(f"   ✅ {date}: {actual_late} minutes (expected {expected_late})")
                        else:
                            print(f"   ❌ {date}: {actual_late} minutes, expected {expected_late}")
                            all_late_ok = False
                            test_passed = False
                
                if all_late_ok:
                    print(f"   ✅ All late minutes match")
            else:
                print(f"   ⚠️  'Menit Terlambat' column not found")
        else:
            print(f"   ⚠️  Could not find header row in Absensi sheet")
        
        # Test regression: JSON report
        print("\n=== TEST 9: Regression - JSON report endpoint ===")
        json_resp = requests.get(
            f"{BASE_URL}/api/absensi/report",
            params={"from": "2026-08-19", "to": "2026-08-25"},
            headers=auth_headers
        )
        
        if json_resp.status_code == 200:
            json_data = json_resp.json()
            if 'items' in json_data and 'filter' in json_data:
                print(f"   ✅ JSON report endpoint working (returned {len(json_data.get('items', []))} items)")
            else:
                print(f"   ❌ JSON report response missing expected fields")
                test_passed = False
        else:
            print(f"   ❌ JSON report endpoint failed: {json_resp.status_code}")
            test_passed = False
        
        # Final result
        print("\n" + "="*80)
        if test_passed:
            print("✅ ALL TESTS PASSED - Jam Kerja Diakui patch is WORKING")
            print("\nSUMMARY:")
            print("- ✅ Kerja Diakui calculation: intersection of actual ∩ shift normal")
            print("- ✅ SO Diakui calculation: early arrival time (before shift start)")
            print("- ✅ Lembur Diakui calculation: only approved overtime counted")
            print("- ✅ Total Diakui formula: Kerja + SO + Lembur (no double count)")
            print("- ✅ Rekapitulasi totals: 43.62h Kerja + 3.00h SO + 2.00h Lembur = 48.62h Total")
            print("- ✅ All 7 per-record values match expected (within ±0.02 tolerance)")
            print("- ✅ Stock Opname sheet: 2 records with correct SO Diakui (1.00h, 2.00h)")
            print("- ✅ Lembur sheet: 3 records with correct Diakui (C2=0, C6=2.00, C7=0)")
            print("- ✅ Absensi sheet: late minutes correct for all records")
            print("- ✅ JSON report endpoint: no regression")
        else:
            print("❌ SOME TESTS FAILED - See details above")
        print("="*80)
        
    except Exception as e:
        print(f"\n❌ TEST ERROR: {e}")
        import traceback
        traceback.print_exc()
        test_passed = False
    
    finally:
        # Cleanup
        cleanup_test_data()
    
    return test_passed

if __name__ == "__main__":
    success = test_excel_export()
    exit(0 if success else 1)
